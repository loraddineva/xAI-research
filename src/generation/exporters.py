"""
src/generation/exporters.py
Writers that persist a generation run to disk in three complementary formats.

Layout produced per run:

    outputs/generation/<run_id>/
        narratives.jsonl   — one record per line; streamed during the run
        run.json           — single self-contained file: full config + all records
        narratives.xlsx    — flat spreadsheet, one row per narrative

Every record includes the full rendered prompt and a `model` block with
provider / model_name / temperature / max_tokens so downstream consumers
(evaluation, manual inspection, paper appendices) can reconstruct exactly
what was sent to the LLM and which configuration produced it.

Public API
----------
    NarrativeRecord          — dataclass capturing one row.
    write_jsonl(path, records)
    write_run_json(path, run_metadata, records)
    write_xlsx(path, records)
    record_to_jsonl_dict(rec) — used by the streaming writer in generator.py.
"""

from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, List, Mapping, Sequence


# ---------------------------------------------------------------------------
# Record dataclass
# ---------------------------------------------------------------------------

@dataclass
class NarrativeRecord:
    """One generated narrative + everything needed to reproduce it."""

    narrative_id: str
    run_id: str
    dataset: str
    instance_id: int
    model_id: str
    model_provider: str
    model_name: str
    temperature: float
    max_tokens: int
    pred_proba: float
    pred_label: int
    shap_values_sorted: List[List[Any]]   # [[feature, value], ...] most positive → most negative
    prompt: str
    narrative_text: str
    created_at: str
    error: str = ""                       # populated only if generation failed

    def to_dict(self) -> dict:
        """Flat dict — convenient for XLSX export."""
        return asdict(self)


# ---------------------------------------------------------------------------
# JSONL helpers
# ---------------------------------------------------------------------------

def record_to_jsonl_dict(rec: NarrativeRecord) -> dict:
    """
    Build the structured (nested) dict written to JSONL / run.json.

    The flat dataclass form is retained for tabular outputs; this nested
    form is closer to what an external consumer (paper appendix, dashboard)
    would expect.
    """
    return {
        "narrative_id": rec.narrative_id,
        "run_id": rec.run_id,
        "dataset": rec.dataset,
        "instance_id": rec.instance_id,
        "model": {
            "id": rec.model_id,
            "provider": rec.model_provider,
            "model_name": rec.model_name,
            "max_tokens": rec.max_tokens,
            "temperature": rec.temperature,
        },
        "prediction": {
            "proba": rec.pred_proba,
            "label": rec.pred_label,
        },
        "shap_values_sorted": rec.shap_values_sorted,
        "prompt": rec.prompt,
        "narrative_text": rec.narrative_text,
        "error": rec.error,
        "created_at": rec.created_at,
    }


def append_jsonl(path: Path, rec: NarrativeRecord) -> None:
    """
    Append a single narrative record as one JSONL line.

    Used by the streaming writer in generator.run_generation so that a
    crash mid-run does not lose previously-generated narratives.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(record_to_jsonl_dict(rec)) + "\n")


def write_jsonl(path: Path, records: Iterable[NarrativeRecord]) -> None:
    """Write all records to a JSONL file (one record per line)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        for rec in records:
            fh.write(json.dumps(record_to_jsonl_dict(rec)) + "\n")


# ---------------------------------------------------------------------------
# Consolidated JSON
# ---------------------------------------------------------------------------

def write_run_json(
    path: Path,
    run_metadata: Mapping[str, Any],
    records: Sequence[NarrativeRecord],
) -> None:
    """
    Write the full run as one JSON file: metadata at the top, all narratives
    nested under "narratives". Suitable for archiving a single run as one
    self-contained, version-controllable artefact.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "run": dict(run_metadata),
        "narratives": [record_to_jsonl_dict(r) for r in records],
    }
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

# Order of columns in the XLSX output. Stable across runs so the file is
# diff-friendly when archived.
_XLSX_COLUMNS: List[str] = [
    "narrative_id",
    "run_id",
    "dataset",
    "instance_id",
    "model_id",
    "model_provider",
    "model_name",
    "temperature",
    "max_tokens",
    "pred_proba",
    "pred_label",
    "shap_values_sorted",
    "prompt",
    "narrative_text",
    "error",
    "created_at",
]


def write_xlsx(path: Path, records: Sequence[NarrativeRecord]) -> None:
    """
    Write all records to a single-sheet XLSX file.

    The `shap_values_sorted` cell is JSON-encoded so the spreadsheet stays
    rectangular (no nested cells).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for XLSX export. "
            "Run: pip install openpyxl"
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "narratives"

    ws.append(_XLSX_COLUMNS)

    for rec in records:
        row = []
        for col in _XLSX_COLUMNS:
            val = getattr(rec, col)
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            row.append(val)
        ws.append(row)

    # Modest column widths — long text cells (prompt, narrative_text) wrap
    # in Excel anyway; this just keeps the metadata columns readable.
    width_hints = {
        "narrative_id": 36,
        "run_id": 30,
        "dataset": 14,
        "instance_id": 12,
        "model_id": 16,
        "model_provider": 14,
        "model_name": 30,
        "temperature": 12,
        "max_tokens": 12,
        "pred_proba": 12,
        "pred_label": 12,
        "shap_values_sorted": 60,
        "prompt": 80,
        "narrative_text": 80,
        "error": 30,
        "created_at": 28,
    }
    for idx, col in enumerate(_XLSX_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_hints.get(col, 18)

    wb.save(path)
